"""Anexos do chat de jurimetria — extração de texto de arquivos enviados.

Espelha o Horizon/smart-mail: o usuário anexa, a mensagem ganha o marcador
[arquivo: <nome> #<uuid>] e o agente lê via tool `ler_arquivo`. Só o TEXTO
extraído é persistido (ChatFile.texto) — binário não fica no banco.

Formatos: PDF (pypdf), XLSX (openpyxl), e texto puro (txt/csv/md/json/xml).
Fail-soft: qualquer falha vira {'erro': ...} — nunca derruba o upload/tool.
"""
from __future__ import annotations

import logging

logger = logging.getLogger(__name__)

MAX_UPLOAD_MB = 15
MAX_TEXTO_CHARS = 300_000  # ~75k tokens — teto do que guardamos por arquivo

_TEXT_EXTS = {'txt', 'csv', 'md', 'json', 'xml', 'html', 'log'}


def _ext(nome: str) -> str:
    return (nome.rsplit('.', 1)[-1] if '.' in nome else '').lower()


def extrair_texto(nome: str, blob: bytes) -> tuple[str, str]:
    """(texto, erro). Erro vazio = sucesso."""
    ext = _ext(nome)
    try:
        if ext == 'pdf':
            import io

            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(blob))
            partes = []
            for pg in reader.pages:
                partes.append(pg.extract_text() or '')
                if sum(len(p) for p in partes) > MAX_TEXTO_CHARS:
                    break
            texto = '\n\n'.join(partes).strip()
            if not texto:
                return '', 'PDF sem texto extraível (provável digitalização/imagem — sem OCR)'
            return texto[:MAX_TEXTO_CHARS], ''
        if ext == 'xlsx':
            import io

            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
            linhas = []
            for ws in wb.worksheets:
                linhas.append(f'## Planilha: {ws.title}')
                for row in ws.iter_rows(values_only=True):
                    linhas.append('\t'.join('' if c is None else str(c) for c in row))
                    if sum(len(x) for x in linhas) > MAX_TEXTO_CHARS:
                        break
            return '\n'.join(linhas)[:MAX_TEXTO_CHARS], ''
        if ext in _TEXT_EXTS:
            return blob.decode('utf-8', errors='replace')[:MAX_TEXTO_CHARS], ''
        return '', f'formato .{ext or "?"} não suportado (aceito: pdf, xlsx, {", ".join(sorted(_TEXT_EXTS))})'
    except Exception as exc:  # noqa: BLE001
        logger.warning('chat_arquivos.extrair_texto %s falhou: %s', nome, exc)
        return '', f'falha ao extrair texto ({type(exc).__name__})'


def ler_arquivo(file_id: str, offset: int = 0, max_chars: int = 8000) -> dict:
    """Tool: lê o texto de um arquivo anexado na conversa, paginado por chars."""
    from .models import ChatFile
    try:
        f = ChatFile.objects.get(uuid=str(file_id).strip().lstrip('#'))
    except Exception:  # noqa: BLE001 — uuid inválido/inexistente
        return {'erro': f'arquivo não encontrado: {file_id}'}
    max_chars = min(max(int(max_chars or 8000), 500), 16000)
    offset = max(int(offset or 0), 0)
    trecho = f.texto[offset:offset + max_chars]
    prox = offset + len(trecho)
    return {'file_id': str(f.uuid), 'filename': f.filename, 'total_chars': f.chars,
            'offset': offset, 'devolvidos': len(trecho),
            'proximo_offset': prox if prox < f.chars else None, 'texto': trecho}
